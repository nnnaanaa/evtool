import os
import re
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Font

# 自然順ソート用関数
def natural_sort_key(s):
    return [int(text) if text.isdigit() else text.lower()
            for text in re.split(r'(\d+)', s)]

# 設定
image_folder = "img"
output_excel = "images.xlsx"
max_images_per_sheet = 20
display_width = 300 * 2.5
display_height = 200 * 2.5

# 対象画像を自然順で取得
image_files = sorted(
    [f for f in os.listdir(image_folder) if f.lower().endswith(('.png', '.jpg', '.jpeg'))],
    key=natural_sort_key
)

# Excel作成
wb = Workbook()
index_ws = wb.active
index_ws.title = "目次"

# 初期化
sheet_links = []
sheet = None

# メイン処理
for i, image_file in enumerate(image_files):
    # シート切替が必要なタイミング
    if i % max_images_per_sheet == 0:
        sheet_num = (i // max_images_per_sheet) + 1
        sheet_name = f"画像一覧{sheet_num}"
        sheet = wb.create_sheet(title=sheet_name)
        sheet.column_dimensions['A'].width = display_width / 7

        # 目次に追加（このシートの最初の画像）
        sheet_links.append((sheet_name, image_file))
        row = 1  # 行リセット

    image_path = os.path.join(image_folder, image_file)

    # キャプション
    cell = sheet.cell(row=row, column=1, value=image_file)
    cell.font = Font(size=14, bold=True)

    # 画像貼り付け
    excel_img = ExcelImage(image_path)
    excel_img.width = display_width
    excel_img.height = display_height
    sheet.add_image(excel_img, f"A{row + 1}")

    # 行の高さ調整
    sheet.row_dimensions[row + 1].height = display_height * 0.75

    # 次の画像の行位置を調整
    # row += int(display_height / 15) + 4
    row += int(4)

# 目次シート作成
for i, (sheet_name, first_image) in enumerate(sheet_links, start=1):
    index_ws.cell(row=i, column=1).value = f"{sheet_name}（{first_image}～）"
    index_ws.cell(row=i, column=1).hyperlink = f"#'{sheet_name}'!A1"
    index_ws.cell(row=i, column=1).style = "Hyperlink"

# 不要なデフォルトシート削除（最初の空シート）
if "Sheet" in wb.sheetnames:
    wb.remove(wb["Sheet"])

# 保存
wb.save(output_excel)
print(f"{output_excel} に画像を正しく20件ずつ分割して貼り付けました。")
